# 데이터베이스 연동과 웹 스크래핑

# 1. SQLite 데이터베이스
print("=== SQLite 데이터베이스 예제 ===")
import sqlite3

# 데이터베이스 연결
conn = sqlite3.connect('example.db')
cursor = conn.cursor()

# 테이블 생성
cursor.execute('''
CREATE TABLE IF NOT EXISTS users (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL,
    email TEXT UNIQUE,
    age INTEGER
)
''')

# 데이터 삽입
try:
    cursor.execute("INSERT INTO users (name, email, age) VALUES (?, ?, ?)",
                  ("홍길동", "hong@example.com", 25))
    cursor.execute("INSERT INTO users (name, email, age) VALUES (?, ?, ?)",
                  ("김철수", "kim@example.com", 30))
    conn.commit()
except sqlite3.IntegrityError:
    print("이미 존재하는 이메일입니다.")

# 데이터 조회
cursor.execute("SELECT * FROM users")
users = cursor.fetchall()
print("\n사용자 목록:")
for user in users:
    print(f"ID: {user[0]}, 이름: {user[1]}, 이메일: {user[2]}, 나이: {user[3]}")

# 데이터베이스 연결 종료
conn.close()

# 2. MySQL 데이터베이스 (mysql-connector-python 필요)
print("\n=== MySQL 데이터베이스 예제 ===")
"""
import mysql.connector

# 데이터베이스 연결
db = mysql.connector.connect(
    host="localhost",
    user="your_username",
    password="your_password",
    database="your_database"
)

cursor = db.cursor()

# 테이블 생성
cursor.execute('''
CREATE TABLE IF NOT EXISTS products (
    id INT AUTO_INCREMENT PRIMARY KEY,
    name VARCHAR(255),
    price DECIMAL(10, 2)
)
''')

# 데이터 삽입
sql = "INSERT INTO products (name, price) VALUES (%s, %s)"
values = ("노트북", 1200000)
cursor.execute(sql, values)
db.commit()

# 데이터 조회
cursor.execute("SELECT * FROM products")
products = cursor.fetchall()
for product in products:
    print(f"ID: {product[0]}, 이름: {product[1]}, 가격: {product[2]}")

db.close()
"""

# 3. 웹 스크래핑 (Beautiful Soup)
print("\n=== 웹 스크래핑 예제 ===")
"""
import requests
from bs4 import BeautifulSoup

# 웹 페이지 요청
url = "https://example.com"
response = requests.get(url)
soup = BeautifulSoup(response.text, 'html.parser')

# HTML 파싱
title = soup.title.text
print(f"페이지 제목: {title}")

# 특정 요소 찾기
links = soup.find_all('a')
print("\n링크 목록:")
for link in links:
    print(f"텍스트: {link.text}, URL: {link.get('href')}")
"""

# 4. Selenium을 이용한 동적 웹 스크래핑
print("\n=== Selenium 웹 스크래핑 예제 ===")
"""
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# 웹드라이버 설정
driver = webdriver.Chrome()

try:
    # 웹페이지 접속
    driver.get("https://example.com")
    
    # 요소 대기 및 클릭
    element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, "myDynamicElement"))
    )
    element.click()
    
    # 결과 가져오기
    results = driver.find_elements(By.CLASS_NAME, "result")
    for result in results:
        print(result.text)

finally:
    driver.quit()
"""

# 5. API 요청 (requests)
print("\n=== API 요청 예제 ===")
"""
import requests

# GET 요청
response = requests.get('https://api.example.com/data')
data = response.json()
print("API 응답:", data)

# POST 요청
payload = {'key': 'value'}
response = requests.post('https://api.example.com/post', json=payload)
print("POST 응답:", response.json())
"""

# 6. 데이터 저장 (CSV)
print("\n=== CSV 파일 처리 예제 ===")
import csv

# CSV 파일 쓰기
with open('data.csv', 'w', newline='', encoding='utf-8') as file:
    writer = csv.writer(file)
    writer.writerow(['이름', '나이', '도시'])
    writer.writerow(['홍길동', 25, '서울'])
    writer.writerow(['김철수', 30, '부산'])

# CSV 파일 읽기
with open('data.csv', 'r', encoding='utf-8') as file:
    reader = csv.reader(file)
    print("\nCSV 데이터:")
    for row in reader:
        print(row)

# 7. Excel 파일 처리 (openpyxl)
print("\n=== Excel 파일 처리 예제 ===")
"""
from openpyxl import Workbook, load_workbook

# Excel 파일 생성
wb = Workbook()
ws = wb.active
ws['A1'] = '이름'
ws['B1'] = '나이'
ws['A2'] = '홍길동'
ws['B2'] = 25
wb.save('data.xlsx')

# Excel 파일 읽기
wb = load_workbook('data.xlsx')
ws = wb.active
for row in ws.iter_rows(values_only=True):
    print(row)
"""

# 8. JSON 파일 처리
print("\n=== JSON 파일 처리 예제 ===")
import json

# JSON 파일 쓰기
data = {
    "users": [
        {"name": "홍길동", "age": 25},
        {"name": "김철수", "age": 30}
    ]
}

with open('data.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

# JSON 파일 읽기
with open('data.json', 'r', encoding='utf-8') as f:
    loaded_data = json.load(f)
    print("\nJSON 데이터:")
    print(json.dumps(loaded_data, ensure_ascii=False, indent=2)) 